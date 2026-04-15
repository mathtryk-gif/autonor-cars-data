#!/usr/bin/env python3
"""
Parser Autonor.xlsx fra Autologik og producerer cars.json i det format,
som Davids dashboard forventer.

Input : Autonor.xlsx (modtaget som mail-vedhæftning hver nat kl. 23:22)
Output: cars.json    (commit'es til repo'et autonor-cars-data)

Dedupliceringslogik:
 - Hver bil identificeres entydigt af sit stelnummer (VIN).
 - Hvis samme VIN optræder flere gange i samme fil, beholdes den første forekomst.
 - Biler uden gyldigt VIN (tom eller < 10 tegn) ignoreres.
"""
from __future__ import annotations
import json
import sys
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook


# Kolonne-indeks (0-baseret) i Autonor.xlsx
COL_VIN           = 0   # A — Stelnr.
COL_MODEL         = 3   # D — Model
COL_FRA           = 4   # E — Fra
COL_TIL           = 5   # F — Til
COL_KLARDATO      = 6   # G — Klardato
COL_KOERSELSDATO  = 7   # H — Kørselsdato
COL_LEVERING      = 9   # J — Forventet leveringsdato
COL_TRACK         = 11  # L — Track and Trace

# Første rækken med data (1-baseret) — header står i række 3
FIRST_DATA_ROW = 4


def parse_excel(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    cars: list[dict] = []
    seen: set[str] = set()

    for row in ws.iter_rows(min_row=FIRST_DATA_ROW, values_only=True):
        vin_raw = row[COL_VIN]
        if not isinstance(vin_raw, str):
            continue
        vin = vin_raw.strip().upper()
        if len(vin) < 10:
            continue
        if vin in seen:
            continue
        seen.add(vin)

        levering = row[COL_LEVERING]
        leveringsdato = (
            levering.strftime("%Y-%m-%d")
            if isinstance(levering, datetime)
            else None
        )

        cars.append(
            {
                "vin": vin,
                "model": (row[COL_MODEL] or "").strip(),
                "fra": (row[COL_FRA] or "").strip(),
                "til": (row[COL_TIL] or "").strip(),
                "leveringsdato": leveringsdato,
                "track": (row[COL_TRACK] or "").strip() or None,
            }
        )

    return cars


def build_payload(cars: list[dict]) -> dict:
    return {
        "lastUpdated": datetime.now().astimezone().isoformat(timespec="seconds"),
        "source": "autologik-mail",
        "count": len(cars),
        "cars": cars,
    }


def main() -> int:
    if len(sys.argv) < 2:
        print("Brug: parse_autonor.py <input.xlsx> [output.json]", file=sys.stderr)
        return 2

    xlsx = Path(sys.argv[1])
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("cars.json")

    cars = parse_excel(xlsx)
    payload = build_payload(cars)

    out.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"OK — {len(cars)} unikke biler skrevet til {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
